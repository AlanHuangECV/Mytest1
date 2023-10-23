import time
import json
import os
# =========Window=========================
import tkinter as tk
from tkinter import messagebox
# from PIL import Image, ImageTk
# =========Excel=======================
import openpyxl
# from openpyxl.styles import Font
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter



# ===========Email=======================
import win32com.client as win32
# ===========API===============================
import requests

# ==========pytest===========================
import pytest

# from selenium import webdriver
# from selenium.webdriver.support.wait import WebDriverWait
# from selenium.webdriver.common.by import By
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.support import expected_conditions as EC
#
# import pyautogui
# import pprint

Current_Status_code = 0

def A1_login_token(url_link,account,password):
    global Current_Status_code
    # API網址
    # url = "http://127.0.0.1:5000/A1_login/"
    url = url_link
    # print(url)
    # time.sleep(200)
    # 標頭
    headers = {
        # "Content-Type": "application/x-www-form-urlencoded",
        # "Host": "account.kkbox.com"
    }
    # 參數
    data = {
        "account": account ,
        "password": password ,
        "url": "https://ecv_qa_sprint2.ecv-platform.com"
    }
    my_session = requests.Session()  # 用session物件可以自動保存及帶出session,不用再自行填寫
    response = my_session.post(url, headers=headers, data=data)
    Current_Status_code = response.status_code
    # print(response)
    # print(Current_Status_code)
    # print(account)
    # print(password)
    # time.sleep(200)
    try:
        if Current_Status_code == 200:  # 正確
            user_token = response.json()["token"]
            refresh_token = response.json()["refresh_token"]
            retDic = {"token":user_token,"refresh_token":refresh_token}
        else:
            retDic = "HTTP error" + f" Status code:{Current_Status_code}"
            # print(f"fail A1_login_token {Current_Status_code}")#########################################
    except:
        retDic = "Unknown error"



    return retDic


def A4_Re_token(url_link,token, Re_token):
    global Current_Status_code
    # API網址
    # url = "http://127.0.0.1:5000/A1_login/"
    url = url_link
    # 標頭
    headers = {
        # "Content-Type": "application/x-www-form-urlencoded",
        "token": token,
        "refresh-token": Re_token
    }
    # 參數
    # data = {
    #     "account": account,
    #     "password": password,
    #     "url": url_link
    # }
    my_session = requests.Session()  # 用session物件可以自動保存及帶出session,不用再自行填寫
    response = my_session.put(url, headers=headers)
    Current_Status_code = response.status_code
    New_token = response.json()["token"]
    # print(response.json())
    # print(New_token)
    # try:
    #     if Current_Status_code == 200:  # 正確
    #         return_code = str(Current_Status_code)
    #     else:
    #         return_code = "HTTP error" + f" Status code:{Current_Status_code}"
    # except:
    #     return_code = "Unknown error"

    # print(f"A4_Re_token {Current_Status_code}")
    return Current_Status_code

def A5_logout(url_link,token):
    global Current_Status_code
    # API網址
    # url = "http://127.0.0.1:5000/A1_login/"
    url = url_link
    # 標頭
    headers = {
        # "Content-Type": "application/x-www-form-urlencoded",
        "token": token
    }
    # 參數
    # data = {
    #     "account": account,
    #     "password": password,
    #     "url": url_link
    # }
    my_session = requests.Session()  # 用session物件可以自動保存及帶出session,不用再自行填寫
    response = my_session.put(url, headers=headers)
    Current_Status_code = response.status_code
    # try:
    #     if Current_Status_code == 200:  # 正確
    #         return_code = str(Current_Status_code)
    #     else:
    #         return_code = "HTTP error" + f" Status code:{Current_Status_code}"
    # except:
    #     return_code = "Unknown error"

    return Current_Status_code


def A6_listUser(url_link,token,query,page,page_size,sort_column,sort_method):
    global Current_Status_code
    # API網址
    # url = "http://127.0.0.1:5000/A1_login/"
    url = url_link
    # 標頭
    headers = {
        # "Content-Type": "application/x-www-form-urlencoded",
        "token": token
    }
    # Request body
    # data = {
    #     "account": account,
    #     "password": password,
    #     "url": url_link
    # }
    # 參數
    params = {
        "query": query,
        "page": page,
        "page_size": page_size,
        "sort_column": sort_column,
        "sort_method": sort_method
    }

    my_session = requests.Session()  # 用session物件可以自動保存及帶出session,不用再自行填寫
    response = my_session.get(url, headers=headers,params=params)
    Current_Status_code = response.status_code
    retDic = {}
    try:
        if Current_Status_code == 200:  # 正確
            retDic = response.json()["item_list"]
            # for list_count in response.json()["item_list"]:
            #     retDic
            # user_id = response.json()["item_list"]["user_id"]
            # account = response.json()["item_list"]["account"]
            # first_name = response.json()["item_list"]["first_name"]
            # last_name = response.json()["item_list"]["last_name"]
            # role_name = response.json()["item_list"]["role_name"]
            # project_name_list = response.json()["item_list"]["project_name_list"]
            # account_status = response.json()["item_list"]["account_status"]
            # retDic = {"token": user_token, "refresh_token": refresh_token}
        else:
            return_code = "HTTP error" + f" Status code:{Current_Status_code}"
    except:
        return_code = "Unknown error"

    return retDic


def A7_getUser(url_link,token,user_id):
    global Current_Status_code
    url = url_link + str(user_id)
    # 標頭
    headers = {
        "token": token
    }
    # Request body
    # data = {
    #     "user_id": user_id
    # }
    # 參數
    # params = {
    #     "user_id": user_id,
    #
    # }

    my_session = requests.Session()  # 用session物件可以自動保存及帶出session,不用再自行填寫
    response = my_session.get(url, headers=headers)
    Current_Status_code = response.status_code
    # print(user_id)
    retDic = {}
    # print(Current_Status_code)
    try:
        if Current_Status_code == 200:  # 正確
            retDic = {
                "user_id": response.json()["user_id"],
                "account": response.json()["account"],
                "role_id": response.json()["role_id"],
                "project_id_list": response.json()["project_id_list"]
            }
        else:
            retDic = "HTTP error" + f" Status code:{Current_Status_code}"
    except:
        retDic = "Unknown error"

    return retDic


def A8_addUser(url_link,token,account,role_id,project_id_list):
    global Current_Status_code
    url = url_link
    # 標頭
    headers = {
        "accept": "application/json",
        "token": token
    }
    # Request body
    data = {
        "account": account,
        "role_id": role_id,
        "project_id_list": project_id_list
    }

    # 參數
    # params = {
    #     "user_id": user_id,
    #
    # }
    # print(url)
    # pprint.pprint(headers)
    # pprint.pprint(data)
    # time.sleep(200)
    my_session = requests.Session()  # 用session物件可以自動保存及帶出session,不用再自行填寫
    response = my_session.post(url, headers=headers, json=data)
    Current_Status_code = response.status_code



    return Current_Status_code

#=================A9_activateUser 不適合自動化測試==且前置要A8_addUser===========================================================
def A9_activateUser(url_link):
    global Current_Status_code
    url = url_link
    # 標頭
    headers = {
        "accept": "application/json",

    }


    my_session = requests.Session()  # 用session物件可以自動保存及帶出session,不用再自行填寫
    response = my_session.get(url, headers=headers)
    Current_Status_code = response.status_code



    return Current_Status_code


def A11_editUser(url_link,token,user_id,role_id,project_id_list):
    global Current_Status_code
    url = url_link + str(user_id)

    # 標頭
    headers = {
        "token": token
    }
    # Request body
    data = {
        "role_id": role_id,
        "project_id_list": project_id_list
    }
    # 參數
    # params = {
    #     "user_id": user_id,
    #
    # }

    my_session = requests.Session()  # 用session物件可以自動保存及帶出session,不用再自行填寫
    response = my_session.put(url, headers=headers,json=data)
    Current_Status_code = response.status_code


    return Current_Status_code



def A12_deleteUser(url_link,token,user_id):
    global Current_Status_code
    url = url_link + str(user_id)

    # 標頭
    headers = {
        "token": token
    }
    # Request body
    # data = {
    #     "role_id": role_id,
    #     "project_id_list": project_id_list
    # }
    # 參數
    # params = {
    #     "user_id": user_id,
    #
    # }

    my_session = requests.Session()  # 用session物件可以自動保存及帶出session,不用再自行填寫
    response = my_session.delete(url, headers=headers)
    Current_Status_code = response.status_code


    return Current_Status_code



def read_json_file(path):
    with open(path, "r", encoding="utf-8") as f:
        str_json_file = f.read()
        if str_json_file == "123":
            return {}

        else:
            return json.loads(str_json_file)


def append_json_file(path, DatetimeTitle, Test_Id, Test_Result, Failure_Desc,HTTP_CODE):
    json_file_dic = read_json_file (path)
    json_file_dic[DatetimeTitle] = {
        "Test ID" : Test_Id,
        "Test Result" : Test_Result,
        "Failure Description": Failure_Desc,
        "HTTP Code" : HTTP_CODE
    }
    with open(path, "w", encoding="utf-8") as f:
        f.write(json.dumps(json_file_dic, ensure_ascii=False, indent=2))

def get_localtime_str():
    localtime = time.localtime()
    result = time.strftime("%Y%m%d%H%M%S", localtime)
    return result

# def get_today_str():
#     result = str(time.strftime("%Y%m%d", time.localtime()))
#     return result

def sheet_title(sheet):
    sheet['A1'].value = "Date/Time"  # 儲存格 A1 內容為 Date/Time
    sheet['B1'].value = "Test ID"  # 儲存格 B1 內容為 Test ID
    sheet['C1'].value = "Test Result" # 儲存格 C1 內容為 Test Result
    sheet['D1'].value = "Failure Description"  # 儲存格 D1 內容為 Failure Description
    sheet['E1'].value = "HTTP Code"  # 儲存格 E1 內容為 HTTP Code


def get_color(result):
    if result == "Pass":
        color = "green"
    else:
        color = "red"

    return color

def show_test_result(Test_Result,Test_Id):
    window = tk.Tk()  # 創造一個視窗
    window.title("Test Result")  # 視窗標題
    window.geometry("900x160")  # 設定視窗初始大小
    set_window_center(window,900, 160)  # 將視窗放在中央
    # window.maxsize(width="1024", height="768") # 設定視窗最大的大小
    window.resizable(False, False)  # 設定視窗大小能不能被改變
    window.config(padx=10, pady=10)  # 調位置,不要擠在左上角
    color = get_color(Test_Result)
    show_result = tk.Label(text=f"{Test_Id}:{Test_Result}", font=("標楷體", 50), width=24, fg=color)  # 設定lable文字
    show_result.pack()
    window.update()
    time.sleep(1)
    window.destroy()


def set_window_center(window,win_width, win_height): # 傳入視窗大小,將視窗設在螢幕中央
    # 获取屏幕 宽、高
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    # 計算 x, y 位置
    x = int(round((screen_width/2) - (win_width/2),0))
    y = int(round((screen_height/2) - (win_height/2),0))
    # window.geometry('%dx%d+%d+%d' % (win_width, win_height, x, y))
    window.geometry(f"{win_width}x{win_height}+{x}+{y}")


def load_test_case(TestCaseFILE):
    global sysPATH
    TestCasePATH = os.path.join(sysPATH,"TestCase", TestCaseFILE)
    if not os.path.isfile(TestCasePATH):
        # print(TestCasePATH)
        # print(f"無{TestCaseFILE}檔,程式將結束")
        messagebox.showerror("File not found",f"無{TestCasePATH}檔,程式將結束")
        exit()
    else:
        with open(TestCasePATH, "r", encoding="utf-8") as file:
            str_json_file = file.read()
            json_file_dic = json.loads(str_json_file)

    return json_file_dic


def A1test():
    TestCaseFILE = "TestCase_API_A1.json"
    json_file_dic = load_test_case(TestCaseFILE)
    Func_url = "/api/v1/user/log_in"
    url = Com_url+Func_url
    # print(url)
    # time.sleep(200)
    for Test_Id in json_file_dic:
        Test_Result = ""
        try:
            user_token = A1_login_token(url, json_file_dic[Test_Id]["ID"], json_file_dic[Test_Id]["password"])["token"]
            # print(Current_Status_code)
            # time.sleep(200)
            if Current_Status_code == 200 :
                err = "錯誤ID/PW可以登入"
                Test_Result = "Fail"
                DatetimeTitle = get_localtime_str()
                sFail.append([DatetimeTitle, Test_Id, Test_Result, err,Current_Status_code])
            elif Current_Status_code == 401 or Current_Status_code == 422 :
                err = ""
                Test_Result = "Pass"
                DatetimeTitle = get_localtime_str()
                sPass.append([DatetimeTitle, Test_Id, Test_Result, err,Current_Status_code])
            elif Current_Status_code == 500 :
                err = "500 Internal Server Error"
                Test_Result = "Fail"
                DatetimeTitle = get_localtime_str()
                sFail.append([DatetimeTitle, Test_Id, Test_Result, err,Current_Status_code])
            else:
                err = "Unknown Error"
                Test_Result = "Fail"
                DatetimeTitle = get_localtime_str()
                sFail.append([DatetimeTitle, Test_Id, Test_Result, err,Current_Status_code])

        except:
            if Current_Status_code == 401 or Current_Status_code == 422 :
                err = ""
                Test_Result = "Pass"
                DatetimeTitle = get_localtime_str()
                sPass.append([DatetimeTitle, Test_Id, Test_Result, err,Current_Status_code])
            else:
                err = "Unknown except Error"
                Test_Result = "Fail"
                DatetimeTitle = get_localtime_str()
                sFail.append([DatetimeTitle, Test_Id, Test_Result, err,Current_Status_code])

        finally:
            DatetimeTitle = get_localtime_str()
            append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err,Current_Status_code)
            sAll.append([DatetimeTitle, Test_Id, Test_Result, err,Current_Status_code])
            time.sleep(0.5)
            wb.save(file_name)  # 儲存檔案
            # wb.close()  # 關閉檔案
            show_test_result(Test_Result,Test_Id)  # 視窗顯示測試結果
            print(Current_Status_code)
#============最後才測正確登入=============================================================================================
    try:
        user_token = A1_login_token(url, user_ID, user_PW)["token"]
        if Current_Status_code == 200:
            err = ""
            Test_Result = "Pass"
            DatetimeTitle = get_localtime_str()
            sPass.append([DatetimeTitle, "A1_user_login", Test_Result, err,Current_Status_code])
        elif Current_Status_code == 401:
            err = "401 使用者身分驗證失敗"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A1_user_login", Test_Result, err,Current_Status_code])
        elif Current_Status_code == 422:
            err = "422 請求參數驗證錯誤"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A1_user_login", Test_Result, err,Current_Status_code])
        elif Current_Status_code == 500:
            err = "500 Internal Server Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A1_user_login", Test_Result, err,Current_Status_code])
        else:
            err = "Unknown Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A1_user_login", Test_Result, err,Current_Status_code])

    except:
        err = "Unknown except Error"
        Test_Result = "Fail"
        DatetimeTitle = get_localtime_str()
        sFail.append([DatetimeTitle, "A1_user_login", Test_Result, err,Current_Status_code])

    finally:
        DatetimeTitle = get_localtime_str()
        append_json_file(TestLogPath, DatetimeTitle, "A1_user_login", Test_Result, err,Current_Status_code)
        sAll.append([DatetimeTitle, "A1_user_login", Test_Result, err,Current_Status_code])
        time.sleep(0.5)
        wb.save(file_name)  # 儲存檔案
        # wb.close()  # 關閉檔案
        show_test_result(Test_Result, "A1_user_login")  # 視窗顯示測試結果
        print(Current_Status_code)
        return Test_Result


def A4test():

    Func_url = "/api/v1/user/refresh_token"
    url = Com_url + Func_url
    token_url = Login_url
    token_Dic = A1_login_token(token_url, user_ID, user_PW)
    user_token = token_Dic["token"]
    refresh_token = token_Dic["refresh_token"]
    # print(user_token)
    # print(refresh_token)
    # print(f"token: {user_token}")
    # ret = A5_logout(url, user_token)

    try:
        ret = A4_Re_token(url, user_token,refresh_token)
        print(ret)
        if ret == 200:
            err = ""
            Test_Result = "Pass"
            DatetimeTitle = get_localtime_str()
            sPass.append([DatetimeTitle, "A4_Re_token", Test_Result, err,Current_Status_code])
        elif ret == 401:
            err = "401 使用者身分驗證失敗"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A4_Re_token", Test_Result, err,Current_Status_code])
        elif ret == 422:
            err = "422 請求參數驗證錯誤"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A4_Re_token", Test_Result, err,Current_Status_code])
        elif ret == 500:
            err = "500 Internal Server Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A4_Re_token", Test_Result, err,Current_Status_code])
        else:
            err = "Unknown Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A4_Re_token", Test_Result, err,Current_Status_code])

    except:
        err = "Unknown Error"
        Test_Result = "Fail"
        DatetimeTitle = get_localtime_str()
        sFail.append([DatetimeTitle, "A4_Re_token", Test_Result, err,Current_Status_code])

    finally:
        DatetimeTitle = get_localtime_str()
        append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err,Current_Status_code)
        sAll.append([DatetimeTitle, "A4_Re_token", Test_Result, err,Current_Status_code])
        time.sleep(0.5)
        wb.save(file_name)  # 儲存檔案
        # wb.close()  # 關閉檔案
        show_test_result(Test_Result, "A4_Re_token")
        print(Current_Status_code)
        return Test_Result


def A5test():

    # TestCaseFILE = "TestCase_API_A1.json"
    # json_file_dic = load_test_case(TestCaseFILE)
    Func_url = "/api/v1/user/log_out"
    url = Com_url + Func_url
    token_url = Login_url
    user_token = A1_login_token(token_url, user_ID, user_PW)["token"]
    # print(f"token: {user_token}")
    # ret = A5_logout(url, user_token)

    try:
        ret = A5_logout(url, user_token)
        if ret == 200:
            err = ""
            Test_Result = "Pass"
            DatetimeTitle = get_localtime_str()
            sPass.append([DatetimeTitle, "A5_user_logout", Test_Result, err,Current_Status_code])
        elif ret == 401:
            err = "401 使用者身分驗證失敗"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A5_user_logout", Test_Result, err,Current_Status_code])
        elif ret == 422:
            err = "422 請求參數驗證錯誤"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A5_user_logout", Test_Result, err,Current_Status_code])
        elif ret == 500:
            err = "500 Internal Server Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A5_user_logout", Test_Result, err,Current_Status_code])
        else:
            err = "Unknown Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, "A5_user_logout", Test_Result, err,Current_Status_code])

    except:
        err = "Unknown Error"
        Test_Result = "Fail"
        DatetimeTitle = get_localtime_str()
        sFail.append([DatetimeTitle, "A5_user_logout", Test_Result, err,Current_Status_code])

    finally:
        DatetimeTitle = get_localtime_str()
        append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err,Current_Status_code)
        sAll.append([DatetimeTitle, "A5_user_logout", Test_Result, err,Current_Status_code])
        time.sleep(0.5)
        wb.save(file_name)  # 儲存檔案
        # wb.close()  # 關閉檔案
        show_test_result(Test_Result, "A5_user_logout")
        print(Current_Status_code)
        return Test_Result

def A6test():
    global First_user_id
    Func_url = "/api/v1/user/list_user"
    url = Com_url + Func_url
    token_url = Login_url
    user_token = A1_login_token(token_url, user_ID, user_PW)["token"]
    API_Name = "A6_listUser"
    # print(f"token: {user_token}")
    # ret = A5_logout(url, user_token)

    try:
        ret = A6_listUser(url, user_token,"",1,10,"account","asc")
        First_user_id = ret[0]["user_id"]
        # print(ret)
        # print(ret[0]["user_id"])
        # time.sleep(200)
        if Current_Status_code == 200:
            err = ""
            Test_Result = "Pass"
            DatetimeTitle = get_localtime_str()
            sPass.append([DatetimeTitle, API_Name, Test_Result, err,Current_Status_code])
        elif Current_Status_code == 401:
            err = "401 使用者身分驗證失敗"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err,Current_Status_code])
        elif Current_Status_code == 422:
            err = "422 請求參數驗證錯誤"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err,Current_Status_code])
        elif Current_Status_code == 500:
            err = "500 Internal Server Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err,Current_Status_code])
        else:
            err = "Unknown Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err,Current_Status_code])

    except:
        err = "Unknown Error"
        Test_Result = "Fail"
        DatetimeTitle = get_localtime_str()
        sFail.append([DatetimeTitle, API_Name, Test_Result, err,Current_Status_code])

    finally:
        DatetimeTitle = get_localtime_str()
        append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err,Current_Status_code)
        sAll.append([DatetimeTitle, API_Name, Test_Result, err,Current_Status_code])
        time.sleep(0.5)
        wb.save(file_name)  # 儲存檔案
        # wb.close()  # 關閉檔案
        show_test_result(Test_Result, API_Name)
        print(Current_Status_code)
        return Test_Result


def A7test():

    Func_url = "/api/v1/user/get_user/"
    url = Com_url + Func_url
    token_url = Login_url
    user_token = A1_login_token(token_url, user_ID, user_PW)["token"]
    API_Name = "A7_getUser"
    # print(f"token: {user_token}")
    # ret = A5_logout(url, user_token)

    try:
        ret = A7_getUser(url, user_token, First_user_id)
        # print(ret)
        # print(ret[0]["user_id"])
        # time.sleep(200)
        # print(f"First_user_id: {First_user_id}\n user_id: {ret['user_id']}")
        if Current_Status_code == 200:
            if First_user_id == ret["user_id"]:
                err = ""
                Test_Result = "Pass"
                DatetimeTitle = get_localtime_str()
                sPass.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
            else:
                err = "User ID mis-match"
                Test_Result = "Fail"
                DatetimeTitle = get_localtime_str()
                sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 401:
            err = "401 使用者身分驗證失敗"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 422:
            err = "422 請求參數驗證錯誤"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 500:
            err = "500 Internal Server Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        else:
            err = "Unknown Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])

    except:
        err = "Unknown Error"
        Test_Result = "Fail"
        DatetimeTitle = get_localtime_str()
        sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])

    finally:
        DatetimeTitle = get_localtime_str()
        append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err, Current_Status_code)
        sAll.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        time.sleep(0.5)
        wb.save(file_name)  # 儲存檔案
        # wb.close()  # 關閉檔案
        show_test_result(Test_Result, API_Name)
        print(Current_Status_code)
        return Test_Result

def A8test():

    Func_url = "/api/v1/user/add_user"
    url = Com_url + Func_url
    token_url = Login_url
    user_token = A1_login_token(token_url, user_ID, user_PW)["token"]
    API_Name = "A8_addUser"
    #==============測試資料建議採用讀檔方式===========================================
    # account = "alan.huang@ecloudvalley.com"
    account = "test9912399 @ abc.com"
    role_id = 2
    project_id_list = [1,2]
    try:
        ret = A8_addUser(url, user_token, account,role_id,project_id_list)
        if Current_Status_code == 200:
            err = ""
            Test_Result = "Pass"
            DatetimeTitle = get_localtime_str()
            sPass.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 401:
            err = "401 使用者身分驗證失敗"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 422:
            err = "422 請求參數驗證錯誤"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 500:
            err = "500 Internal Server Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        else:
            err = "Unknown Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])

    except:
        err = "Unknown Error"
        Test_Result = "Fail"
        DatetimeTitle = get_localtime_str()
        sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])

    finally:
        DatetimeTitle = get_localtime_str()
        append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err, Current_Status_code)
        sAll.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        time.sleep(0.5)
        wb.save(file_name)  # 儲存檔案
        # wb.close()  # 關閉檔案
        show_test_result(Test_Result, API_Name)
        print(Current_Status_code)
        return Test_Result

#============A9test不適合自動化測試,且須跟A8一起測========================================================================
def A9test():

    Func_url = "/api/v1/user/activate_user"
    Activate_token = "eyJhbGciOiAiUlMyNTYiLCAidHlwIjogIkpXVCJ9.eyJlbWFpbCI6ICJhbGFuLmh1YW5nQGVjbG91ZHZhbGxleS5jb20iLCAiaG9zdG5hbWUiOiAiZWN2X3FhX3NwcmludDIuZWN2b3QuY29tIiwgImV4cCI6IDE2OTc3Njk0NjIuNDA5NX0.EYU78M6RttxTry_SM8y8k-k1eCViy7CG8WCT8UPlq4aXWWuPsK5govJ1W4rlkbvyddLAhjT8LRkNue51yssGBsPrcLs6PzD5HJBO0HAWeDv2Y01D_Jjcia_vDvphCFipPkswTL8tw08vAjtQ3W8rLS1EPvIfByDq2_FMj25c3B6yVW2mnWrRjPMsvTk9c_rkmzx4ivKYkrkUBsRy5FTKIRAK0_DdJEWfK9TCgIAg2CbJAfg4tJjZ-gxE2Oe0XR-Bj6gVfWM_3DbXAWdLRDWq3VLm8m8UDdLbgl-57r0bck7TXDDpxylKQQi5qj94mGX7vI0_UdEDhyiMt-yN0Gy1_Q"
    url = Com_url + Func_url + "?token=" + Activate_token
    API_Name = "A9_activateUser"
    # ==============測試資料建議採用讀檔方式===========================================

    try:
        ret = A9_activateUser(url)
        if Current_Status_code == 200:
            err = ""
            Test_Result = "Pass"
            DatetimeTitle = get_localtime_str()
            sPass.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 401:
            err = "401 使用者身分驗證失敗"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 422:
            err = "422 請求參數驗證錯誤"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 500:
            err = "500 Internal Server Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        else:
            err = "Unknown Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])

    except:
        err = "Unknown Error"
        Test_Result = "Fail"
        DatetimeTitle = get_localtime_str()
        sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])

    finally:
        DatetimeTitle = get_localtime_str()
        append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err, Current_Status_code)
        sAll.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        time.sleep(0.5)
        wb.save(file_name)  # 儲存檔案
        # wb.close()  # 關閉檔案
        show_test_result(Test_Result, API_Name)
        print(Current_Status_code)
        return Test_Result
def A11test():

    Func_url = "/api/v1/user/edit_user/"
    url = Com_url + Func_url
    token_url = Login_url
    user_token = A1_login_token(token_url, user_ID, user_PW)["token"]
    API_Name = "A11_editUser"
    A6_url= Com_url+"/api/v1/user/list_user"
    ret_listUser= A6_listUser(A6_url, user_token, "", 1, 10, "account", "asc")
    First_user_id = ret_listUser[0]["user_id"]
    role_id=2
    project_id_list = [1,2]

    try:
        ret = A11_editUser(url, user_token, First_user_id ,role_id,project_id_list)
        if Current_Status_code == 200:
            err = ""
            Test_Result = "Pass"
            DatetimeTitle = get_localtime_str()
            sPass.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 401:
            err = "401 使用者身分驗證失敗"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 422:
            err = "422 請求參數驗證錯誤"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 500:
            err = "500 Internal Server Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        else:
            err = "Unknown Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])

    except:
        err = "Unknown Error"
        Test_Result = "Fail"
        DatetimeTitle = get_localtime_str()
        sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])

    finally:
        DatetimeTitle = get_localtime_str()
        append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err, Current_Status_code)
        sAll.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        time.sleep(0.5)
        wb.save(file_name)  # 儲存檔案
        # wb.close()  # 關閉檔案
        show_test_result(Test_Result, API_Name)
        print(Current_Status_code)
        return Test_Result

#=================注意:A12test為危險測試必須搭配A8_addUser服用==========================================================================
def A12test():

    Func_url = "/api/v1/user/delete_user/"
    url = Com_url + Func_url
    token_url = Login_url
    user_token = A1_login_token(token_url, user_ID, user_PW)["token"]
    API_Name = "A12_deleteUser"
    A8_url = Com_url + "/api/v1/user/add_user"
    account = "test9912399@abc.com"
    role_id = 2
    project_id_list = [1, 2]
    ret_addUser = A8_addUser(A8_url, user_token, account, role_id, project_id_list)
    # print(Current_Status_code)
    # time.sleep(200)
    if Current_Status_code == 200:
        A6_url = Com_url + "/api/v1/user/list_user"
        ret_listUser = A6_listUser(A6_url, user_token, "", 1, 10000, "account", "asc")
        # pprint.pprint(ret_listUser[0]["user_id"])
        # time.sleep(200)
        for i in range(len(ret_listUser)):
            if ret_listUser[i]["account"] == account:
                user_id = ret_listUser[i]["user_id"]
    else:
        err = "Add user Error"
        Test_Result = "Fail"
        DatetimeTitle = get_localtime_str()
        sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])

    # print(user_id)
    # time.sleep(200)
    try:
        ret = A12_deleteUser(url, user_token,user_id)
        if Current_Status_code == 200:
            err = ""
            Test_Result = "Pass"
            DatetimeTitle = get_localtime_str()
            sPass.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 401:
            err = "401 使用者身分驗證失敗"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 422:
            err = "422 請求參數驗證錯誤"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        elif Current_Status_code == 500:
            err = "500 Internal Server Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        else:
            err = "Unknown Error"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])

    except:
        err = "Unknown Error"
        Test_Result = "Fail"
        DatetimeTitle = get_localtime_str()
        sFail.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])

    finally:
        DatetimeTitle = get_localtime_str()
        append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err, Current_Status_code)
        sAll.append([DatetimeTitle, API_Name, Test_Result, err, Current_Status_code])
        time.sleep(0.5)
        wb.save(file_name)  # 儲存檔案
        # wb.close()  # 關閉檔案
        show_test_result(Test_Result, API_Name)
        print(Current_Status_code)
        return Test_Result





    # =================視窗顯示測試結果==================================
    # while True:
    #     print(pyautogui.position()) #找座標
    #     time.sleep(0.3)
    # ====================end====================================

def send_email(subject, body,HTMLBody, recipients,cc,attachedPATH):
    try:
        outlook = win32.Dispatch('Outlook.Application')
        mail = outlook.CreateItem(0)
        mail.Subject = subject
        # mail.Body = body
        # mail.HTMLBody = '''<html><head><meta http-equiv="Content-Type" content="text/html; charset=utf-8" /><meta http-equiv="Content-Style-Type" content="text/css" /><meta name="generator" content="Aspose.Words for .NET 17.1.0.0" /><title></title></head><body><div><p style="margin-top:0pt; margin-bottom:0pt; widows:0; orphans:0; font-size:12pt"><span style="font-family:'等线'; font-weight:bold">請注意</span><span style="font-family:Calibri; font-weight:bold">: </span><span style="font-family:'等线'; font-weight:bold">附件</span><span style="font-family:'等线'; font-weight:bold">僅為測試結果</span><span style="font-family:'等线'; font-weight:bold">並非</span><span style="font-family:Calibri; font-weight:bold">”</span><span style="font-family:Calibri; font-weight:bold">Bug</span><span style="font-family:Calibri; font-weight:bold">”</span><span style="font-family:Calibri; font-weight:bold">,</span><span style="font-family:'等线'; font-weight:bold; color:#0070c0">僅供參考</span><span style="font-family:Calibri; font-weight:bold">.</span><span style="font-family:Calibri; font-weight:bold"> </span><span style="font-family:'等线'; font-weight:bold">尚</span><span style="font-family:'等线'; font-weight:bold">需</span><span style="font-family:'等线'; font-weight:bold">要</span><span style="font-family:Calibri; font-weight:bold">Q</span><span style="font-family:Calibri; font-weight:bold">A</span><span style="font-family:'等线'; font-weight:bold">釐</span><span style="font-family:'等线'; font-weight:bold">清及</span><span style="font-family:'等线'; font-weight:bold">確認</span></p></div></body></html>'''
        mail.HTMLBody = HTMLBody
        mail.To = recipients
        mail.CC = cc  # 指定抄送人
        mail.Attachments.Add(attachedPATH)
        # mail.Attachments.Add(attachedPATH, 1, 1, "myFile")  # 例子添加附件
        # mail.save()  # 保存，换.send()可以发送
        mail.Send()
        print("郵件寄出成功！")

    except Exception as e:
        # print("郵件寄出失敗:", str(e))
        messagebox.showerror("錯誤", f"郵件寄出失敗: {e}")
        # show_test_result("郵件寄出失敗:", "")



def clear_entry():
    entry_fast.delete(0,tk.END)
    entry_integrated.delete(0,tk.END)
    entry_All.delete(0, tk.END)

def ExeTest():
    global GUI_window

    GUI_window.minsize()
    fast_times = entry_fast.get() # A1 login test
    integrated_times = entry_integrated.get()
    All_times = entry_All.get()
    if fast_times == "":
        fast_times = 0
    elif fast_times.isnumeric() == False:
        messagebox.showwarning("警告","請輸入正整數!")
        return
    else:
        fast_times = int (fast_times)

    if integrated_times == "":
        integrated_times = 0
    elif integrated_times.isnumeric() == False:
        messagebox.showwarning("警告","請輸入正整數!")
        return
    else:
        integrated_times = int (integrated_times)

    if All_times == "":
        All_times = 0
    elif All_times.isnumeric() == False:
        messagebox.showwarning("警告","請輸入正整數!")
        return
    else:
        All_times = int (All_times)

    GUI_window.destroy()

    while fast_times > 0:

        A1test()
        A4test()  # A4測試
        A5test()  # A5測試
        A6test()  # A7要跟A6一起測
        A7test()  # A7要跟A6一起測
        A8test()
        A11test()
        fast_times -=1

    while integrated_times > 0:
        A12test()
        integrated_times -=1

    while All_times > 0:

        A1test()  # A1測試
        A4test()  # A4測試
        A5test()  # A5測試
        A6test()  # A7要跟A6一起測
        A7test()  # A7要跟A6一起測
        A8test()
        A11test()
        All_times -=1

#==============API begin================================================================
# url = "http://127.0.0.1:5000/A1_login/"
# user_token = A1_login_token(url,"ABC","123")
# # print(user_token)
#
# url = "http://127.0.0.1:5000/A5_logout/"
# ret = A5logout(url,user_token)
# print(ret)
# # time.sleep(200)

#==============API end================================================================

# ProgramName = "DeviceOn"
# TargetWebsite = "https://deviceon-trial.wise-paas.com/"
Com_url = "https://ne-user-beta.ecvot.com"
Login_url = Com_url+"/api/v1/user/log_in"
First_user_id = 0
user_ID = ""
user_PW = ""
# DeviceOn_ID_PW =[]
PATH = "Login.json"
PATH = os.path.join(os.getcwd(),PATH)
# TestLogPath = "TestLog.json"
Test_Id = ""
TIMEOUT = 2
sysPATH = os.getcwd()
Input_username = ""
Input_password = ""

# =============設定allure==============================================
if __name__ == '__main__':
    # pytest.main([r'--alluredir=test_report\allure'])
    pytest.main(["-s", "-v", "--html=Outputs/reports/pytest.html", "--alluredir=Outputs/allure"])   # allure文件生成的目錄



today_str = str(time.strftime("%Y%m%d", time.localtime()))
TestLogPath = "API_TestLog" + today_str + ".json"

LogPath = os.path.join(os.getcwd(),"Logs") # 路徑為程式所在的"Log資料夾"
if not os.path.isdir(LogPath): # 如果沒有就建立
    os.mkdir(LogPath)

os.chdir(r""+LogPath) # 'r'是不轉義字符
file_name = "".join([today_str,".xlsx"]) # 合併字串
if not os.path.isfile(file_name):
    wb = openpyxl.Workbook()    # 建立空白的 Excel 活頁簿物件
    sPass = wb['Sheet']
    sPass.title = "Pass"
    sPass.sheet_properties.tabColor = "00ff00"
    sFail = wb.create_sheet("Fail")
    sFail.sheet_properties.tabColor = "ff0000"
    sAll = wb.create_sheet("All")
    sAll.sheet_properties.tabColor = "000000"
    sheet_title(sPass)
    sheet_title(sFail)
    sheet_title(sAll)
    wb.save(file_name)       # 儲存檔案
    wb.close()  # 關閉檔案

else:
    wb = openpyxl.load_workbook(file_name) # 不要開來開去會無法寫入
    # wb = openpyxl.load_workbook(file_name, data_only=True)  # 不要開來開去會無法寫入
    sPass = wb["Pass"]
    sFail = wb["Fail"]
    sAll = wb["All"]
    try:
        wb.save(file_name)  # 儲存檔案
    except IOError as e:

        # 创建 Excel 应用程序对象
        excel = win32.Dispatch("Excel.Application")

        # 关闭文件
        excel.Workbooks.Close()

        # 退出 Excel 应用程序
        excel.Quit()
    except Exception as e:
        messagebox.showerror("錯誤", f"Test report 建立失敗: {e}")
        messagebox.showwarning("警告", "程式即將結束")
        exit()




sPass.column_dimensions['A'].width = 17
sPass.column_dimensions['B'].width = 13
sPass.column_dimensions['C'].width = 11
sPass.column_dimensions['D'].width = 30
sPass.column_dimensions['E'].width = 11
sFail.column_dimensions['A'].width = 17
sFail.column_dimensions['B'].width = 13
sFail.column_dimensions['C'].width = 11
sFail.column_dimensions['D'].width = 30
sFail.column_dimensions['E'].width = 11
sAll.column_dimensions['A'].width = 17
sAll.column_dimensions['B'].width = 13
sAll.column_dimensions['C'].width = 11
sAll.column_dimensions['D'].width = 30
sAll.column_dimensions['E'].width = 11


if not os.path.isfile(TestLogPath):
    with open(TestLogPath, "w", encoding="utf-8") as f: #先建空檔案寫123, 之後會蓋掉
        f.write(json.dumps(123, ensure_ascii=False))

# ===========先讀ini參數=============================================================================
if not os.path.isfile(PATH):
    print("無Password檔,程式將結束")
    exit()
else:
    with open(PATH, "r", encoding= "utf-8") as file:  # 讀ini檔
        str_json_file = file.read() # 先讀json檔, 但此時為字串string
        json_file_dic = json.loads(str_json_file) # 將讀到的字串string改為字典dictionary格式

user_ID = json_file_dic["Login"]["ID"]
user_PW = json_file_dic["Login"]["password"]

assert A1test() == "Pass"
assert A4test() == "Pass"
assert A5test() == "Pass"
assert A6test() == "Pass"  # A7要跟A6一起測
assert A7test() == "Pass"  # A7要跟A6一起測
assert A8test() == "Pass"
assert A11test() == "Pass"
assert A12test() == "Pass"
# ===================完成讀取參數===================================================
# ===================#開始生成待測網址視窗 begin=======================================
# browser = webdriver.Chrome()
# browser.maximize_window() # 縮到最小
# ===================#測試登入目標網址===================================================
# #############===================#開始生成GUI介面 begin===================================================================
# GUI_window = tk.Tk() # 創造一個視窗
# GUI_window.title("API test") # 視窗標題
# GUI_window.geometry("500x420") # 設定視窗初始大小
# set_window_center(GUI_window,500,420) # 將視窗放在中央
# # window.maxsize(width="1024", height="768") # 設定視窗最大的大小
# # GUI_window.resizable(False,False) # 設定視窗大小能不能被改變
# GUI_window.config(padx=10,pady=10) # 調位置,不要擠在左上角
# Img = Image.open(os.path.join(sysPATH,"Pic/NeoEdge.png"))        # 開啟圖片
# # Img = Image.open("Pic/NeoEdge.png")        # 開啟圖片
# tk_img = ImageTk.PhotoImage(Img) # 轉換為 tk 圖片物件
# GUI_window.iconphoto(True,tk_img) # 改icon圖片: 第一個參數True是設定新視窗是否沿用同圖片
# # tk.Toplevel() # 新增視窗
# # 設定背景顏色(若去背),若window背景顏色一樣會有白邊, 用'highlightthickness=0'去白邊
# canvas = tk.Canvas(GUI_window, width=480, height=151, highlightthickness=0) # NeoEdge圖片為480x151像素
# # 在 Canvas 中放入圖片, 錨點在西北,如果不放錨點前2參數是圖片中心點
# canvas.create_image(0, 0, image=tk_img,anchor='nw')# 在 Canvas 中放入圖片, 錨點在西北
# canvas.grid(row=0, column=0, columnspan=3,pady =5) # 長度佔2個column
# # canvas.pack()
#
# # =======================排位置==================================================
# label_fast = tk.Label(text= "Done test:" ,font = ("標楷體", 16),width=10, anchor="w") # 設定lable文字
# label_fast.grid(row=1, column = 0) # 設定label出現位置,否則不會出現
# entry_fast = tk.Entry(font = ("標楷體", 16), fg = "green",width=8)
# entry_fast.grid(row=1, column = 1) # 設定Entry出現位置,否則不會出現
# label_bout1 = tk.Label(text= "次" ,font = ("標楷體", 16),width=6, anchor="e") # 設定lable文字
# label_bout1.grid(row=1, column = 2) # 設定label出現位置,否則不會出現
# label_integrated = tk.Label(text= "New test:" ,font = ("標楷體", 16),width=10, anchor="w") # 設定lable文字
# label_integrated.grid(row=2, column = 0, pady=5) # 設定label出現位置,否則不會出現
# entry_integrated = tk.Entry(font = ("標楷體", 16), fg = "green",width=8)
# entry_integrated.grid(row=2, column = 1) # 設定Entry出現位置,否則不會出現
# label_bout2 = tk.Label(text= "次" ,font = ("標楷體", 16),width=6, anchor="e") # 設定lable文字
# label_bout2.grid(row=2, column = 2) # 設定label出現位置,否則不會出現
# label_All = tk.Label(text= "全部測試:" ,font = ("標楷體", 16),width=10, anchor="w") # 設定lable文字
# label_All.grid(row=3, column = 0) # 設定label出現位置,否則不會出現
# entry_All = tk.Entry(font = ("標楷體", 16), fg = "green",width=8)
# entry_All.grid(row=3, column = 1) # 設定Entry出現位置,否則不會出現
# label_bout3 = tk.Label(text= "次" ,font = ("標楷體", 16),width=6, anchor="e") # 設定lable文字
# label_bout3.grid(row=3, column = 2) # 設定label出現位置,否則不會出現
# button_Execute = tk.Button(text= "執行" ,font = ("標楷體", 16),width=30,fg="white",bg="#00BB00", command= ExeTest) # 設定按鈕文字
# button_Execute.grid(row=4, column = 0, pady=10, columnspan=3)  # 設定按鈕出現位置,否則不會出現
# button_Clear = tk.Button(text= "清除" ,font = ("標楷體", 16),width=30,fg="white",bg="red", command= clear_entry) # 設定按鈕文字
# button_Clear.grid(row=5, column = 0, pady=10, columnspan=3)  # 設定按鈕出現位置,否則不會出現
# # label = tk.Label(window, image=tk_img, width=200, height=200)  # 在 Lable 中放入圖片
# # label.pack()
# GUI_window.mainloop() # 讓視窗一直存在,偵測事件的無窮迴圈
# #######===================#開始生成GUI介面 end=================================================================================


# ============================================================================






# window.mainloop()
# window.quit()





attachedPATH = os.path.join(os.getcwd(),file_name) # 路徑為程式所在的"Log資料夾"
# 設置郵件主题、正文和收件人
subject = "QA team Auto-mail, do NOT reply !"
# body = "測試自動寄信的郵件。"
body = ""
HTMLBody = '''<html><head><meta http-equiv="Content-Type" content="text/html; charset=utf-8" /><meta http-equiv="Content-Style-Type" content="text/css" /><meta name="generator" content="Aspose.Words for .NET 17.1.0.0" /><title></title></head><body><div><p style="margin-top:0pt; margin-bottom:0pt; widows:0; orphans:0; font-size:12pt"><span style="font-family:'等线'; font-weight:bold">請注意</span><span style="font-family:Calibri; font-weight:bold">: </span><span style="font-family:'等线'; font-weight:bold">附件</span><span style="font-family:'等线'; font-weight:bold">僅為測試結果</span><span style="font-family:'等线'; font-weight:bold">並非</span><span style="font-family:Calibri; font-weight:bold">”Bug”</span><span style="font-family:Calibri; font-weight:bold">,</span><span style="font-family:'等线'; font-weight:bold; color:#0070c0">僅供參考</span><span style="font-family:Calibri; font-weight:bold">.</span><span style="font-family:Calibri; font-weight:bold"> </span><span style="font-family:'等线'; font-weight:bold">尚</span><span style="font-family:'等线'; font-weight:bold">需</span><span style="font-family:'等线'; font-weight:bold">要</span><span style="font-family:Calibri; font-weight:bold">Q</span><span style="font-family:Calibri; font-weight:bold">A</span><span style="font-family:'等线'; font-weight:bold">釐清及</span><span style="font-family:'等线'; font-weight:bold">確認</span></p><p style="margin-top:0pt; margin-bottom:0pt; widows:0; orphans:0; font-size:12pt"><span style="font-family:Calibri; font-weight:bold">Note: The </span><span style="font-family:Calibri; font-weight:bold">attached file was just test result</span><span style="font-family:Calibri; font-weight:bold">, instead of confirmed “Bug”</span><span style="font-family:Calibri; font-weight:bold">, </span><span style="font-family:Calibri; font-weight:bold; color:#0070c0">just for your reference</span><span style="font-family:Calibri; font-weight:bold">.</span></p><p style="margin-top:0pt; margin-bottom:0pt; widows:0; orphans:0; font-size:12pt"><span style="font-family:Calibri; font-weight:bold">S</span><span style="font-family:Calibri; font-weight:bold">till need clarif</span><span style="font-family:Calibri; font-weight:bold">ication</span><span style="font-family:Calibri; font-weight:bold"> and confirm</span><span style="font-family:Calibri; font-weight:bold">ation by </span><span style="font-family:Calibri; font-weight:bold">QA team</span><span style="font-family:Calibri; font-weight:bold">.</span></p></div></body></html>'''
recipients = "alan.huang@ecloudvalley.com;alan.huang@allianzvantage.com"
cclist = "alan.huang@ecloudvalley.com"

# 寄出郵件
# send_email(subject, body,HTMLBody, recipients,cclist,attachedPATH)#===========vvvvvvvvvv


# retValue = os.system(file_name) # 開啟Excel檔測試結果
TestResultFile = os.popen(file_name) # 開啟Excel檔測試結果 ===========vvvvvvvvvv
exit()


