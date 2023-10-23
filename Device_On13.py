import time
import json

from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
import os
import pyautogui
# =========Window=========================
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
# =========Excel=======================
import openpyxl
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# ===========Email=======================
import win32com.client as win32







def read_json_file(path):
    with open(path, "r", encoding="utf-8") as f:
        str_json_file = f.read()
        if str_json_file == "123":
            return {}

        else:
            return json.loads(str_json_file)


def append_json_file(path, DatetimeTitle, Test_Id, Test_Result, Failure_Desc):
    json_file_dic = read_json_file (path)
    json_file_dic[DatetimeTitle] = {
        "Test ID" : Test_Id,
        "Test Result" : Test_Result,
        "Failure Description": Failure_Desc
    }
    with open(path, "w", encoding="utf-8") as f:
        f.write(json.dumps(json_file_dic, ensure_ascii=False, indent=2))

def get_localtime_str():
    localtime = time.localtime()
    result = time.strftime("%Y%m%d%H%M%S", localtime)
    return result

# def get_today_str():
#     result = str(time.strftime("%Y%m%d", time.localtime()))
#     return result

def sheet_title(sheet):
    sheet['A1'].value = "Date/Time"  # 儲存格 A1 內容為 Date/Time
    sheet['B1'].value = "Test ID"  # 儲存格 B1 內容為 Test ID
    sheet['C1'].value = "Test Result" # 儲存格 C1 內容為 Test Result
    sheet['D1'].value = "Failure Description"  # 儲存格 D1 內容為 Failure Description


def get_color(result):
    if result == "Pass":
        color = "green"
    else:
        color = "red"

    return color

def show_test_result(Test_Result,Test_Id):
    window = tk.Tk()  # 創造一個視窗
    window.title("Test Result")  # 視窗標題
    window.geometry("740x160")  # 設定視窗初始大小
    set_window_center(window,740, 160)  # 將視窗放在中央
    # window.maxsize(width="1024", height="768") # 設定視窗最大的大小
    window.resizable(False, False)  # 設定視窗大小能不能被改變
    window.config(padx=10, pady=10)  # 調位置,不要擠在左上角
    color = get_color(Test_Result)
    show_result = tk.Label(text=f"{Test_Id}:{Test_Result}", font=("標楷體", 50), width=24, fg=color)  # 設定lable文字
    show_result.pack()
    window.update()
    time.sleep(1)
    window.destroy()


def set_window_center(window,win_width, win_height): # 傳入視窗大小,將視窗設在螢幕中央
    # 获取屏幕 宽、高
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    # 計算 x, y 位置
    x = int(round((screen_width/2) - (win_width/2),0))
    y = int(round((screen_height/2) - (win_height/2),0))
    # window.geometry('%dx%d+%d+%d' % (win_width, win_height, x, y))
    window.geometry(f"{win_width}x{win_height}+{x}+{y}")


def load_test_case(TestCaseFILE):
    global sysPATH
    TestCasePATH = os.path.join(sysPATH,"TestCase", TestCaseFILE)
    if not os.path.isfile(TestCasePATH):
        # print(TestCasePATH)
        # print(f"無{TestCaseFILE}檔,程式將結束")
        messagebox.showerror("File not found",f"無{TestCasePATH}檔,程式將結束")
        exit()
    else:
        with open(TestCasePATH, "r", encoding="utf-8") as file:
            str_json_file = file.read()
            json_file_dic = json.loads(str_json_file)

    return json_file_dic


def A002test():
    TestCaseFILE = "TestCaseA002.json"
    json_file_dic = load_test_case(TestCaseFILE)

    for Test_Id in json_file_dic:
        Test_Result = ""
        Input_username.clear()
        # browser.execute_script("arguments[0].focus();", Input_username) # 設焦點在元素上
        browser.execute_script("arguments[0].value = '';", Input_username)  # 清除元素上文字,很重要(Input_username.clear()指令失效)
        time.sleep(0.2)
        Input_username.send_keys(json_file_dic[Test_Id]["ID"])
        Input_password.clear()
        # browser.execute_script("arguments[0].focus();", Input_password)  # 設焦點在元素上
        browser.execute_script("arguments[0].value = '';", Input_password)  # 清除元素上文字,很重要(Input_username.clear()指令失效)
        time.sleep(0.2)
        Input_password.send_keys(json_file_dic[Test_Id]["password"])
        time.sleep(0.2)
        try:
            warn_msg = WebDriverWait(browser, timeout=TIMEOUT).until(
                EC.presence_of_element_located((By.XPATH, '// small[@ class = "text-danger"] ')))
            # print("找到了")
        except:
            err = "警告訊息沒顯示"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, Test_Id, Test_Result, err])
        else:
            # print("如果try執行完都沒問題就過來這邊")
            try:
                # Login_button = WebDriverWait(browser, timeout=10).until(EC.presence_of_element_located(
                #     (By.XPATH, '//*[@id="app"]/div/div/div/div[1]/div/div[2]/div[3]/button')))
                Login_button = WebDriverWait(browser, timeout=TIMEOUT).until(EC.presence_of_element_located(
                    (By.XPATH, '//button[contains(@ class, "solidButton")]')))  # 找到有enable的按鍵
            except:
                err = ""
                Test_Result = "Pass"
                DatetimeTitle = get_localtime_str()
                sPass.append([DatetimeTitle, Test_Id, Test_Result, err])
            else:
                err = "登入按鍵不應該顯示"
                Test_Result = "Fail"
                DatetimeTitle = get_localtime_str()
                sFail.append([DatetimeTitle, Test_Id, Test_Result, err])
        finally:
            DatetimeTitle = get_localtime_str()
            append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err)
            sAll.append([DatetimeTitle, Test_Id, Test_Result, err])
            time.sleep(0.5)
            wb.save(file_name)  # 儲存檔案
            # wb.close()  # 關閉檔案
            show_test_result(Test_Result,Test_Id)  # 視窗顯示測試結果

def A003test():
    TestCaseFILE = "TestCaseA003.json"
    json_file_dic = load_test_case(TestCaseFILE)

    for Test_Id in json_file_dic:
        Test_Result = ""
        Input_username.clear()
        # browser.execute_script("arguments[0].focus();", Input_username) # 設焦點在元素上
        browser.execute_script("arguments[0].value = '';", Input_username)  # 清除元素上文字,很重要(Input_username.clear()指令失效)
        time.sleep(0.2)
        Input_username.send_keys(json_file_dic[Test_Id]["ID"])
        Input_password.clear()
        # browser.execute_script("arguments[0].focus();", Input_password)  # 設焦點在元素上
        browser.execute_script("arguments[0].value = '';", Input_password)  # 清除元素上文字,很重要(Input_username.clear()指令失效)
        time.sleep(0.2)
        Input_password.send_keys(json_file_dic[Test_Id]["password"])
        time.sleep(0.2)
        try:
            Login_button = WebDriverWait(browser, timeout=TIMEOUT).until(EC.presence_of_element_located(
                (By.XPATH, '//button[contains(@ class, "solidButton")]')))  # 找到有enable的按鍵
        except:
            err = "登入按鈕沒顯示"
            Test_Result = "Fail"
            DatetimeTitle = get_localtime_str()
            sFail.append([DatetimeTitle, Test_Id, Test_Result, err])
        else:
            # print("如果try執行完都沒問題就過來這邊")
            Login_button.click()
            try:
                err_confirm_button = WebDriverWait(browser, timeout=TIMEOUT).until(EC.presence_of_element_located(
                    (By.XPATH, '//button[contains(@ class, "swal2-confirm")]')))  # 找到"確認"的按鍵
            except:
                err = "無顯示錯誤確認"
                Test_Result = "Fail"
                DatetimeTitle = get_localtime_str()
                sFail.append([DatetimeTitle, Test_Id, Test_Result, err])
            else:
                err = ""
                Test_Result = "Pass"
                DatetimeTitle = get_localtime_str()
                sPass.append([DatetimeTitle, Test_Id, Test_Result, err])
                err_confirm_button.click()
        finally:
            DatetimeTitle = get_localtime_str()
            append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err)
            sAll.append([DatetimeTitle, Test_Id, Test_Result, err])
            time.sleep(0.5)
            wb.save(file_name)  # 儲存檔案
            # wb.close()  # 關閉檔案
            show_test_result(Test_Result, Test_Id)  # 視窗顯示測試結果

def DefaulTest(TargetWebsite):
    Test_Id = "A001"
    global Input_username
    global Input_password
    global DefaultTestDone
    try:
        browser.get(TargetWebsite)
    except Exception as err:
        Test_Result = "Fail"
        DatetimeTitle = get_localtime_str()
        sFail.append([DatetimeTitle, Test_Id, Test_Result, err])
    else:
        # print("如果try執行完都沒問題就過來這邊")
        err = ""
        Test_Result = "Pass"
        DatetimeTitle = get_localtime_str()
        sPass.append([DatetimeTitle, Test_Id, Test_Result, err])
    finally:
        # print("無論有沒有錯都過來這邊")

        DatetimeTitle = get_localtime_str()
        append_json_file(TestLogPath, DatetimeTitle, Test_Id, Test_Result, err)
        sAll.append([DatetimeTitle, Test_Id, Test_Result, err])
        time.sleep(2)
        wb.save(file_name)  # 儲存檔案

    Input_username = WebDriverWait(browser, timeout=TIMEOUT).until(EC.presence_of_element_located((By.NAME, "txtUserName")))
    Input_password = WebDriverWait(browser, timeout=TIMEOUT).until(EC.presence_of_element_located((By.ID, "txtPasswd")))
    DefaultTestDone = True
    show_test_result(Test_Result, Test_Id)

    # =================視窗顯示測試結果==================================
    # while True:
    #     print(pyautogui.position()) #找座標
    #     time.sleep(0.3)
    # ====================end====================================

def send_email(subject, body,HTMLBody, recipients,cc,attachedPATH):
    try:
        outlook = win32.Dispatch('Outlook.Application')
        mail = outlook.CreateItem(0)
        mail.Subject = subject
        # mail.Body = body
        # mail.HTMLBody = '''<html><head><meta http-equiv="Content-Type" content="text/html; charset=utf-8" /><meta http-equiv="Content-Style-Type" content="text/css" /><meta name="generator" content="Aspose.Words for .NET 17.1.0.0" /><title></title></head><body><div><p style="margin-top:0pt; margin-bottom:0pt; widows:0; orphans:0; font-size:12pt"><span style="font-family:'等线'; font-weight:bold">請注意</span><span style="font-family:Calibri; font-weight:bold">: </span><span style="font-family:'等线'; font-weight:bold">附件</span><span style="font-family:'等线'; font-weight:bold">僅為測試結果</span><span style="font-family:'等线'; font-weight:bold">並非</span><span style="font-family:Calibri; font-weight:bold">”</span><span style="font-family:Calibri; font-weight:bold">Bug</span><span style="font-family:Calibri; font-weight:bold">”</span><span style="font-family:Calibri; font-weight:bold">,</span><span style="font-family:'等线'; font-weight:bold; color:#0070c0">僅供參考</span><span style="font-family:Calibri; font-weight:bold">.</span><span style="font-family:Calibri; font-weight:bold"> </span><span style="font-family:'等线'; font-weight:bold">尚</span><span style="font-family:'等线'; font-weight:bold">需</span><span style="font-family:'等线'; font-weight:bold">要</span><span style="font-family:Calibri; font-weight:bold">Q</span><span style="font-family:Calibri; font-weight:bold">A</span><span style="font-family:'等线'; font-weight:bold">釐</span><span style="font-family:'等线'; font-weight:bold">清及</span><span style="font-family:'等线'; font-weight:bold">確認</span></p></div></body></html>'''
        mail.HTMLBody = HTMLBody
        mail.To = recipients
        mail.CC = cc  # 指定抄送人
        mail.Attachments.Add(attachedPATH)
        # mail.Attachments.Add(attachedPATH, 1, 1, "myFile")  # 例子添加附件
        # mail.save()  # 保存，换.send()可以发送
        mail.Send()
        print("郵件寄出成功！")

    except Exception as e:
        # print("郵件寄出失敗:", str(e))
        messagebox.showerror("錯誤", f"郵件寄出失敗: {e}")
        # show_test_result("郵件寄出失敗:", "")



def clear_entry():
    entry_fast.delete(0,tk.END)
    entry_integrated.delete(0,tk.END)
    entry_All.delete(0, tk.END)

def ExeTest():
    global GUI_window
    global TargetWebsite
    GUI_window.minsize()
    fast_times = entry_fast.get()
    integrated_times = entry_integrated.get()
    All_times = entry_All.get()
    if fast_times == "":
        fast_times = 0
    elif fast_times.isnumeric() == False:
        messagebox.showwarning("警告","請輸入正整數!")
        return
    else:
        fast_times = int (fast_times)

    if integrated_times == "":
        integrated_times = 0
    elif integrated_times.isnumeric() == False:
        messagebox.showwarning("警告","請輸入正整數!")
        return
    else:
        integrated_times = int (integrated_times)

    if All_times == "":
        All_times = 0
    elif All_times.isnumeric() == False:
        messagebox.showwarning("警告","請輸入正整數!")
        return
    else:
        All_times = int (All_times)

    GUI_window.destroy()

    while fast_times > 0:
        if not DefaultTestDone:
            DefaulTest(TargetWebsite)

        A002test()  # A002測試
        fast_times -=1

    while integrated_times > 0:
        if not DefaultTestDone:
            DefaulTest(TargetWebsite)
        A003test()  # A003測試
        integrated_times -=1

    while All_times > 0:
        if not DefaultTestDone:
            DefaulTest(TargetWebsite)
        A002test()  # A002測試
        A003test()  # A003測試
        All_times -=1



ProgramName = "DeviceOn"
TargetWebsite = "https://deviceon-trial.wise-paas.com/"
DeviceOn_user_ID = ""
DeviceOn_user_PW = ""
DeviceOn_ID_PW =[]
PATH = "Password.json"
PATH = os.path.join(os.getcwd(),PATH)
# TestLogPath = "TestLog.json"
Test_Id = ""
TIMEOUT = 2
sysPATH = os.getcwd()
Input_username = ""
Input_password = ""
DefaultTestDone = False

today_str = str(time.strftime("%Y%m%d", time.localtime()))
TestLogPath = "TestLog" + today_str + ".json"

LogPath = os.path.join(os.getcwd(),"Logs") # 路徑為程式所在的"Log資料夾"
if not os.path.isdir(LogPath): # 如果沒有就建立
    os.mkdir(LogPath)

os.chdir(r""+LogPath) # 'r'是不轉義字符
file_name = "".join([today_str,".xlsx"]) # 合併字串
if not os.path.isfile(file_name):
    wb = openpyxl.Workbook()    # 建立空白的 Excel 活頁簿物件
    sPass = wb['Sheet']
    sPass.title = "Pass"
    sPass.sheet_properties.tabColor = "00ff00"
    sFail = wb.create_sheet("Fail")
    sFail.sheet_properties.tabColor = "ff0000"
    sAll = wb.create_sheet("All")
    sAll.sheet_properties.tabColor = "000000"
    sheet_title(sPass)
    sheet_title(sFail)
    sheet_title(sAll)
    wb.save(file_name)       # 儲存檔案
    wb.close()  # 關閉檔案

else:
    wb = openpyxl.load_workbook(file_name) # 不要開來開去會無法寫入
    # wb = openpyxl.load_workbook(file_name, data_only=True)  # 不要開來開去會無法寫入
    sPass = wb["Pass"]
    sFail = wb["Fail"]
    sAll = wb["All"]
    try:
        wb.save(file_name)  # 儲存檔案
    except IOError as e:

        # 创建 Excel 应用程序对象
        excel = win32.Dispatch("Excel.Application")

        # 关闭文件
        excel.Workbooks.Close()

        # 退出 Excel 应用程序
        excel.Quit()
    except Exception as e:
        messagebox.showerror("錯誤", f"Test report 建立失敗: {e}")
        messagebox.showwarning("警告", "程式即將結束")
        exit()




sPass.column_dimensions['A'].width = 17
sPass.column_dimensions['B'].width = 13
sPass.column_dimensions['C'].width = 11
sPass.column_dimensions['D'].width = 30
sFail.column_dimensions['A'].width = 17
sFail.column_dimensions['B'].width = 13
sFail.column_dimensions['C'].width = 11
sFail.column_dimensions['D'].width = 30
sAll.column_dimensions['A'].width = 17
sAll.column_dimensions['B'].width = 13
sAll.column_dimensions['C'].width = 11
sAll.column_dimensions['D'].width = 30


if not os.path.isfile(TestLogPath):
    with open(TestLogPath, "w", encoding="utf-8") as f: #先建空檔案寫123, 之後會蓋掉
        f.write(json.dumps(123, ensure_ascii=False))

# ===========先讀ini參數=============================================================================
if not os.path.isfile(PATH):
    print("無Password檔,程式將結束")
    exit()
else:
    with open(PATH, "r", encoding= "utf-8") as file:  # 讀ini檔
        str_json_file = file.read() # 先讀json檔, 但此時為字串string
        json_file_dic = json.loads(str_json_file) # 將讀到的字串string改為字典dictionary格式

DeviceOn_user_ID = json_file_dic[ProgramName]["ID"]
DeviceOn_user_PW = json_file_dic[ProgramName]["password"]
# ===================完成讀取參數===================================================
# ===================#開始生成待測網址視窗 begin=======================================
browser = webdriver.Chrome()
browser.maximize_window() # 縮到最小
# ===================#測試登入目標網址===================================================
# #############===================#開始生成GUI介面 begin===================================================================
GUI_window = tk.Tk() # 創造一個視窗
GUI_window.title("帳號管理工具") # 視窗標題
GUI_window.geometry("500x420") # 設定視窗初始大小
set_window_center(GUI_window,500,420) # 將視窗放在中央
# window.maxsize(width="1024", height="768") # 設定視窗最大的大小
# GUI_window.resizable(False,False) # 設定視窗大小能不能被改變
GUI_window.config(padx=10,pady=10) # 調位置,不要擠在左上角
Img = Image.open(os.path.join(sysPATH,"Pic/NeoEdge.png"))        # 開啟圖片
# Img = Image.open("Pic/NeoEdge.png")        # 開啟圖片
tk_img = ImageTk.PhotoImage(Img) # 轉換為 tk 圖片物件
GUI_window.iconphoto(True,tk_img) # 改icon圖片: 第一個參數True是設定新視窗是否沿用同圖片
# tk.Toplevel() # 新增視窗
# 設定背景顏色(若去背),若window背景顏色一樣會有白邊, 用'highlightthickness=0'去白邊
canvas = tk.Canvas(GUI_window, width=480, height=151, highlightthickness=0) # NeoEdge圖片為480x151像素
# 在 Canvas 中放入圖片, 錨點在西北,如果不放錨點前2參數是圖片中心點
canvas.create_image(0, 0, image=tk_img,anchor='nw')# 在 Canvas 中放入圖片, 錨點在西北
canvas.grid(row=0, column=0, columnspan=3,pady =5) # 長度佔2個column
# canvas.pack()

# =======================排位置==================================================
label_fast = tk.Label(text= "快速測試 :" ,font = ("標楷體", 16),width=10, anchor="w") # 設定lable文字
label_fast.grid(row=1, column = 0) # 設定label出現位置,否則不會出現
entry_fast = tk.Entry(font = ("標楷體", 16), fg = "green",width=8)
entry_fast.grid(row=1, column = 1) # 設定Entry出現位置,否則不會出現
label_bout1 = tk.Label(text= "次" ,font = ("標楷體", 16),width=6, anchor="e") # 設定lable文字
label_bout1.grid(row=1, column = 2) # 設定label出現位置,否則不會出現
label_integrated = tk.Label(text= "整合測試 :" ,font = ("標楷體", 16),width=10, anchor="w") # 設定lable文字
label_integrated.grid(row=2, column = 0, pady=5) # 設定label出現位置,否則不會出現
entry_integrated = tk.Entry(font = ("標楷體", 16), fg = "green",width=8)
entry_integrated.grid(row=2, column = 1) # 設定Entry出現位置,否則不會出現
label_bout2 = tk.Label(text= "次" ,font = ("標楷體", 16),width=6, anchor="e") # 設定lable文字
label_bout2.grid(row=2, column = 2) # 設定label出現位置,否則不會出現
label_All = tk.Label(text= "全部測試 :" ,font = ("標楷體", 16),width=10, anchor="w") # 設定lable文字
label_All.grid(row=3, column = 0) # 設定label出現位置,否則不會出現
entry_All = tk.Entry(font = ("標楷體", 16), fg = "green",width=8)
entry_All.grid(row=3, column = 1) # 設定Entry出現位置,否則不會出現
label_bout3 = tk.Label(text= "次" ,font = ("標楷體", 16),width=6, anchor="e") # 設定lable文字
label_bout3.grid(row=3, column = 2) # 設定label出現位置,否則不會出現
button_Execute = tk.Button(text= "執行" ,font = ("標楷體", 16),width=30,fg="white",bg="#00BB00", command= ExeTest) # 設定按鈕文字
button_Execute.grid(row=4, column = 0, pady=10, columnspan=3)  # 設定按鈕出現位置,否則不會出現
button_Clear = tk.Button(text= "清除" ,font = ("標楷體", 16),width=30,fg="white",bg="red", command= clear_entry) # 設定按鈕文字
button_Clear.grid(row=5, column = 0, pady=10, columnspan=3)  # 設定按鈕出現位置,否則不會出現
# label = tk.Label(window, image=tk_img, width=200, height=200)  # 在 Lable 中放入圖片
# label.pack()
GUI_window.mainloop() # 讓視窗一直存在,偵測事件的無窮迴圈
# #######===================#開始生成GUI介面 end=================================================================================

# Login登入

Input_username.clear()
browser.execute_script("arguments[0].value = '';", Input_username) # 清除元素上文字,很重要(Input_username.clear()指令失效)
time.sleep(0.2)
Input_username.send_keys(DeviceOn_user_ID)
Input_password.clear()
browser.execute_script("arguments[0].value = '';", Input_password) # 清除元素上文字,很重要(Input_username.clear()指令失效)
Input_password.send_keys(DeviceOn_user_PW)
time.sleep(0.2)
Login_button = WebDriverWait(browser, timeout=TIMEOUT).until(EC.presence_of_element_located((By.XPATH, '//button[contains(@ class, "solidButton")]')))
Login_button.click()
# ============================================================================

time.sleep(2)
pyautogui.moveTo(1323,261,1)
time.sleep(1)
pyautogui.click()
# 設備Group
Equip_G = WebDriverWait(browser, timeout=TIMEOUT).until(EC.presence_of_element_located((By.XPATH, '//*[@id="app"]/div/div[1]/aside/div/div/div[1]/ul/div[2]/li/a/span/span[2]')))
Equip_G.click()
time.sleep(1)
# 設備列表
Equip_list = WebDriverWait(browser, timeout=TIMEOUT).until(EC.presence_of_element_located((By.XPATH, '//*[@id="app"]/div/div[1]/aside/div/div/div[1]/ul/div[2]/li/ul/div[1]/li/a/span')))
Equip_list.click()
time.sleep(1)

# 設備監控
Equip_monitor = WebDriverWait(browser, timeout=10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="app"]/div/div[1]/aside/div/div/div[1]/ul/div[2]/li/ul/div[2]/li/a/span')))
Equip_monitor.click()



# window.mainloop()
# window.quit()
try:
    browser.close()
    browser.quit()
except:
    # dos命令杀死chromedriver进程即可
    os.system("taskkill /F /im chromedriver.exe")




attachedPATH = os.path.join(os.getcwd(),file_name) # 路徑為程式所在的"Log資料夾"
# 設置郵件主题、正文和收件人
subject = "QA team Auto-mail, do NOT reply !"
# body = "測試自動寄信的郵件。"
body = ""
HTMLBody = '''<html><head><meta http-equiv="Content-Type" content="text/html; charset=utf-8" /><meta http-equiv="Content-Style-Type" content="text/css" /><meta name="generator" content="Aspose.Words for .NET 17.1.0.0" /><title></title></head><body><div><p style="margin-top:0pt; margin-bottom:0pt; widows:0; orphans:0; font-size:12pt"><span style="font-family:'等线'; font-weight:bold">請注意</span><span style="font-family:Calibri; font-weight:bold">: </span><span style="font-family:'等线'; font-weight:bold">附件</span><span style="font-family:'等线'; font-weight:bold">僅為測試結果</span><span style="font-family:'等线'; font-weight:bold">並非</span><span style="font-family:Calibri; font-weight:bold">”Bug”</span><span style="font-family:Calibri; font-weight:bold">,</span><span style="font-family:'等线'; font-weight:bold; color:#0070c0">僅供參考</span><span style="font-family:Calibri; font-weight:bold">.</span><span style="font-family:Calibri; font-weight:bold"> </span><span style="font-family:'等线'; font-weight:bold">尚</span><span style="font-family:'等线'; font-weight:bold">需</span><span style="font-family:'等线'; font-weight:bold">要</span><span style="font-family:Calibri; font-weight:bold">Q</span><span style="font-family:Calibri; font-weight:bold">A</span><span style="font-family:'等线'; font-weight:bold">釐清及</span><span style="font-family:'等线'; font-weight:bold">確認</span></p><p style="margin-top:0pt; margin-bottom:0pt; widows:0; orphans:0; font-size:12pt"><span style="font-family:Calibri; font-weight:bold">Note: The </span><span style="font-family:Calibri; font-weight:bold">attached file was just test result</span><span style="font-family:Calibri; font-weight:bold">, instead of confirmed “Bug”</span><span style="font-family:Calibri; font-weight:bold">, </span><span style="font-family:Calibri; font-weight:bold; color:#0070c0">just for your reference</span><span style="font-family:Calibri; font-weight:bold">.</span></p><p style="margin-top:0pt; margin-bottom:0pt; widows:0; orphans:0; font-size:12pt"><span style="font-family:Calibri; font-weight:bold">S</span><span style="font-family:Calibri; font-weight:bold">till need clarif</span><span style="font-family:Calibri; font-weight:bold">ication</span><span style="font-family:Calibri; font-weight:bold"> and confirm</span><span style="font-family:Calibri; font-weight:bold">ation by </span><span style="font-family:Calibri; font-weight:bold">QA team</span><span style="font-family:Calibri; font-weight:bold">.</span></p></div></body></html>'''
recipients = "alan.huang@ecloudvalley.com;alan.huang@allianzvantage.com"
cclist = "alan.huang@ecloudvalley.com"

# 寄出郵件
send_email(subject, body,HTMLBody, recipients,cclist,attachedPATH)


# retValue = os.system(file_name) # 開啟Excel檔測試結果
TestResultFile = os.popen(file_name) # 開啟Excel檔測試結果
exit()


